import openpyxl
import json
from pprint import pprint

def get_menu():
    menu = []
    load_wb = openpyxl.load_workbook("../DB_cell/menu_data.xlsx", data_only = True)
    load_ws = load_wb['menu']
    for row in load_ws.iter_rows(min_row=2):
        menu.append([row[0].value,row[1].value])
    load_wb2 = openpyxl.load_workbook("../DB_cell/extra_menu_data.xlsx", data_only = True)
    load_ws2 = load_wb2['menu']
    n = len(menu)
    for row in load_ws2.iter_rows(min_row=2):
        menu.append([row[0].value + n,row[1].value])
    #pprint(menu)
    return menu

def get_restaurant():
    data = {}
    filename = ['설입','신촌1','신촌2','신촌3','상도동1','상도동2','상도동3']
    for i in filename:
        with open ('../Datasets/'+i+'.json','r',encoding='utf-8') as f:
            temp = json.load(f)
            data.update(temp)
            f.close()
    #pprint(data)
    return data

def numbering_restaurant(restaurant):
    t=0
    for i in restaurant.keys():
        t+=1
        restaurant[i].append(['번호',t])
    #pprint(restaurant)
    return restaurant


def pairing_restaurant_menu(restaurant, menu):
    restaurant_menu = []
    for i in restaurant.keys():
        for l in restaurant[i]:
            if l[0] == "메뉴":
                if l[1] != "메뉴 정보 없음":
                    for j in l[1].keys():
                        for m in menu:
                            if j == m[1]:
                                restaurant_menu.append([restaurant[i][7][1],m[0]])
    pprint(restaurant_menu)
    return restaurant_menu

def extract_restaurant_menu(restaurant_menu):
    write_wb = openpyxl.Workbook()
    write_ws = write_wb.create_sheet('restaurant_menu')

    write_ws = write_wb.active
    write_ws = write_wb['restaurant_menu']
    write_ws.append(['restaurant_id','menu_id'])
    for i in restaurant_menu:
        write_ws.append(i)
    write_wb.save('../DB_cell/restaurant_menu_data.xlsx')

if __name__ == "__main__":
    menu = get_menu()
    restaurant = get_restaurant()
    restaurant = numbering_restaurant(restaurant)
    restaurant_menu = pairing_restaurant_menu(restaurant, menu)
    extract_restaurant_menu(restaurant_menu)
