from openpyxl import Workbook
import pandas as pd
from encodings import utf_8
import json
import re

def get_menu(filename):
    data = {}
    with open ('../Datasets/top_menus/'+filename+'대표메뉴.json','r',encoding='utf-8') as f:
        data = json.load(f)
        f.close()
    return data

def get_restaurant(filename):
    data = {}
    with open ('../Datasets/'+filename+'.json','r',encoding='utf-8') as f:
        data = json.load(f)
        f.close()
    return data

def get_periodTime(dict_of_res):
    for i in dict_of_res.keys():
        for j in dict_of_res[i]:
            if j[0]=='영업 시간' and j[1]!='시간 정보 없음':
                if len(j[1]) > 1:
                    if j[1][1].find('브레이크') == -1:
                        print(j[1][1])

def extract_menu(dict_of_menu):
    menu_dict = {}
    menu_list = list(dict_of_menu.values())
    menu_set = set(menu_list)
    menu_list = list(menu_set)

    write_wb = Workbook()
    write_ws = write_wb.create_sheet('menu')

    write_ws = write_wb.active
    write_ws = write_wb['menu']
    write_ws.append(['id','name','isSoup','isSpicy','isSweet','isHot','isMeat','isNoodle','isRice','isBread'])
    t=0
    for i in menu_list:
        t += 1
        menu_dict[i] = t
        write_ws.append([t,i])

    write_wb.save('../DB_cell/menu_data.xlsx')
    return menu_dict

def extract_address(dict_of_res):
    address_dict = {}
    address_list = []
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('address')

    write_ws = write_wb.active
    write_ws = write_wb['address']
    write_ws.append(['id','city','district','road','building_num','floor','location'])

    for i in dict_of_res.keys():
        for j in dict_of_res[i]:
            if j[0] == "주소":
                address_list.append(j[1])

    address_set = set(address_list)
    address_list = list(address_set)

    t=0
    for j in address_list:
        t += 1
        temp = j.split()
        if j.find('상도동') != -1:
            temp.insert(4,'숭입')
        if j.find('봉천동') != -1:
            temp.insert(4,'설입')
        if j.find('창천동') != -1:
            temp.insert(4,'신촌')
        for u in temp:
            if u.find('층') != -1:
                temp[5] = u.strip(',')
                break;
            else:
                temp[5] = 'null'
        a = temp[4]
        temp[4] = temp[5]
        temp[5] = a
        #print(str(t) + ' ' + temp[0] + ' ' + temp[1] + ' ' + temp[2] + ' ' + temp[3].strip(',') + ' ' + temp[4] + ' ' + temp[5])
        address_dict[j] = t
        write_ws.append([t, temp[0], temp[1], temp[2], temp[3].strip(','), temp[4], temp[5]])
    write_wb.save('../DB_cell/address_data.xlsx')
    return address_dict

def extract_restaurant(dict_of_res, menu_dict, address_dict):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('restaurant')

    write_ws = write_wb.active
    write_ws = write_wb['restaurant']
    write_ws.append(['id','name','external_star','address_id','business_date','starttime','fintime','bestmenu_id','franchise','internal_star','restaurant_type'])
    t=0
    for i in dict_of_res.keys():
        t += 1
        rest_list = []
        rest_list.append(t)
        rest_list.append(i[4:])
        for j in dict_of_res[i]:
            if j[0] == "평점":
                rest_list.append(j[1])
            elif j[0] == "대표메뉴":
                if j[1] == '대표메뉴 없음':
                    rest_list.append('null')
                else:
                    rest_list.append(menu_dict[j[1]])
            elif j[0] == "주소":
                rest_list.append(address_dict[j[1]])
            elif j[0] == "프랜차이즈":
                if j[1] == 'True':
                    rest_list.append(1)
                else:
                    rest_list.append(0)
            elif j[0] == '영업 시간':
                day = '월,화,수,목,금,토,일'
                if j[1] == '시간 정보 없음':
                    rest_list.append('null')
                elif j[1][0].find('매일') != -1:
                    rest_list.append(day)
                elif j[1][0].split()[0].find('~') != -1:
                    #print(j[1][0][0] + '~' + j[1][0][2:3])
                    #print(day[day.find(j[1][0][0]):day.find(j[1][0][2:3])+1])
                    rest_list.append(day[day.find(j[1][0][0]):day.find(j[1][0][2:3])+1])
                elif j[1][0].find(',') != -1:
                    rest_list.append(j[1][0].split()[0])
                else:
                    rest_list.append(day)

                if j[1] == '시간 정보 없음':
                    rest_list.append('null')
                    rest_list.append('null')
                else:
                    temp = re.findall('[0-9]+:[0-9]+', j[1][0])
                    #print(j[1][0])
                    #print(temp)
                    rest_list.append(temp[0])
                    if len(temp) == 1:
                        rest_list.append('null')
                    else:
                        rest_list.append(temp[1])
                    #rest_list.append(temp[1])
                    #rest_list.append(temp[3])

        write_ws.append(rest_list)
        print(rest_list)

    write_wb.save('../DB_cell/restaurant_data.xlsx')


if __name__ == "__main__":
    file = ['설입','신촌1','신촌2','신촌3','상도동1','상도동2','상도동3']
    restaurant_data = {}
    menu_data = {}

    for f in file:
        update_restaurant = get_restaurant(f)
        restaurant_data.update(update_restaurant)
        update_menu = get_menu(f)
        menu_data.update(update_menu)

    #get_periodTime(restaurant_data)
    menu_dict = extract_menu(menu_data)
    address_dict = extract_address(restaurant_data)
    extract_restaurant(restaurant_data, menu_dict, address_dict)


# 레스토랑 id, name, franchise , external_star, starttime, fintime, bestmenu_id, restaurant_type, address_id
# 메뉴 id, name
# 주소 restaurant, city, district, road, building_num, floor, location
