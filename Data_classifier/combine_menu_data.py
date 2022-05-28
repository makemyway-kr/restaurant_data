import openpyxl
import json
from pprint import pprint

def get_menu():
    menu = []
    load_wb = openpyxl.load_workbook("../DB_cell/menu_data.xlsx", data_only = True)
    load_ws = load_wb['menu2']
    for row in load_ws.iter_rows(min_row=1):
        menu.append([row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value,row[9].value])
    load_wb2 = openpyxl.load_workbook("../DB_cell/extra_menu_data.xlsx", data_only = True)
    load_ws2 = load_wb2['menu2']
    n = len(menu)
    for row in load_ws2.iter_rows(min_row=2):
        menu.append([row[0].value + n-1,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value,row[9].value])
    pprint(menu)
    return menu

def extract_all_menu(menu):
    write_wb = openpyxl.Workbook()
    write_ws = write_wb.create_sheet('menu')

    write_ws = write_wb.active
    write_ws = write_wb['menu']
    for i in menu:
        write_ws.append(i)
    write_wb.save('../DB_cell/all_menu_data.xlsx')

if __name__ == "__main__":
    menu = get_menu()
    extract_all_menu(menu)
