import openpyxl
from pprint import pprint

def get_keyword_charac():
    load_wb = openpyxl.load_workbook('../DB_cell/menu_key.xlsx', data_only=True)
    load_ws = load_wb['menu_key']
    menu_keys = {}
    for row in load_ws.iter_rows(min_row=2):
        key_charac = []
        key_charac.append(['isSoup',row[1].value])
        key_charac.append(['isSpicy',row[2].value])
        key_charac.append(['isSweet',row[3].value])
        key_charac.append(['isHot',row[4].value])
        key_charac.append(['isMeat',row[5].value])
        key_charac.append(['isNoodle',row[6].value])
        key_charac.append(['isRice',row[7].value])
        key_charac.append(['isBread',row[8].value])
        menu_keys[row[0].value] = key_charac
    #pprint(menu_keys)
    return menu_keys

def get_menu():
    load_wb = openpyxl.load_workbook('../DB_cell/menu_data.xlsx', data_only=True)
    load_ws = load_wb['menu']
    menu = []
    for row in load_ws['B']:
        menu.append(row.value)
    menu.pop(0)
    return menu;

def fill_charac(menu, menu_keys):
    menu_charac = {}
    for m in menu:
        has_key = 0
        charac = []
        charac2 = [['isSoup',0],['isSpicy',0],['isSweet',0],['isHot',0],['isMeat',0],['isNoodle',0],['isRice',0],['isBread',0]]
        for k in menu_keys.keys():
            if m.find(k) != -1:
                has_key = 1
                for i in menu_keys[k]:
                    if i[1] == 1:
                        charac.append(i[0])
                charac_s = set(charac)
                charac = list(charac_s)
                #print(charac)
        #if has_key == False:
            #print(m)
        for i in charac2:
            for j in charac:
                if i[0] == j:
                    i[1] = 1
        menu_charac[m] = charac2
    pprint(menu_charac)
    return menu_charac

def extract_menu(menu_charac):
    write_wb = openpyxl.load_workbook('../DB_cell/menu_data.xlsx', data_only=True)
    write_ws = write_wb.create_sheet('menu2')

    write_ws = write_wb.active
    write_ws = write_wb['menu2']
    write_ws.append(['id','name','isSoup','isSpicy','isSweet','isHot','isMeat','isNoodle','isRice','isBread'])

    t=0
    for m in menu_charac.keys():
        t += 1
        print(str(t) + ' ' + m + ' ' + str(menu_charac[m][0][1])+ ' ' + str(menu_charac[m][1][1])+ ' ' + str(menu_charac[m][2][1])+ ' ' + str(menu_charac[m][3][1])+ ' ' + str(menu_charac[m][4][1])+ ' ' + str(menu_charac[m][5][1])+ ' ' + str(menu_charac[m][6][1]) + ' ' + str(menu_charac[m][7][1]))
        write_ws.append([t,m,menu_charac[m][0][1],menu_charac[m][1][1],menu_charac[m][2][1],menu_charac[m][3][1],menu_charac[m][4][1],menu_charac[m][5][1],menu_charac[m][6][1],menu_charac[m][7][1]])

    write_wb.save('../DB_cell/menu_data.xlsx')


if __name__ == "__main__":
    menu_keys = get_keyword_charac()
    menu = get_menu()
    menu_charac = fill_charac(menu, menu_keys)
    extract_menu(menu_charac)
