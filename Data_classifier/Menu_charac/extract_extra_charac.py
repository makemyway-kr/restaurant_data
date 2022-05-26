from openpyxl import Workbook
import json
from pprint import pprint

def filtering_top_menu():
    allmenus = []
    topmenus = []
    with open('./allmenus.json','r',encoding = 'utf-8') as f:
        allmenus = json.load(f)
        f.close()
    all_menus_set = set(allmenus)
    allmenus = list(all_menus_set)
    with open('./topmenus.json','r',encoding = 'utf-8') as f:
        topmenus = json.load(f)
        f.close()
    top_menus_set = set(topmenus)
    topmenus = list(top_menus_set)
    for a in allmenus:
        for t in topmenus:
            if a == t:
                allmenus.remove(a)
    return allmenus

def get_extra_key():
    extra_key = []
    filename = ['result2_1.json','result2_2.json','result2_3.json']
    for i in filename:
        with open('./data/' + i, 'r', encoding = 'utf_8') as f:
            temp = json.load(f)
            extra_key += temp[1]
            f.close()
    #pprint(extra_key)
    return extra_key

def extract_all_menu(extra_menus):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('menu')

    write_ws = write_wb.active
    write_ws = write_wb['menu']
    write_ws.append(['id','name','isSoup','isSpicy','isSweet','isHot','isMeat','isNoodle','isRice','isBread'])

    t=0
    for i in extra_menus:
        t += 1
        write_ws.append([t,i])

    write_wb.save('../../DB_cell/extra_menu_data.xlsx')

def extract_extra_key(extra_keys):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('menu_key')

    write_ws = write_wb.active
    write_ws = write_wb['menu_key']
    write_ws.append(['name','isSoup','isSpicy','isSweet','isHot','isMeat','isNoodle','isRice','isBread'])

    for i in extra_keys:
        write_ws.append([i])

    write_wb.save('../../DB_cell/extra_menu_key.xlsx')

if __name__ == "__main__":
    extra_menus = filtering_top_menu()
    extra_keys = get_extra_key()
    extract_all_menu(extra_menus)
    extract_extra_key(extra_keys)
