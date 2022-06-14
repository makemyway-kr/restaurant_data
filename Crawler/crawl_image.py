import openpyxl
import os
import sys
import time
import json
from urllib.parse import quote_plus
from urllib.request import urlopen
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import multiprocessing
from pprint import pprint

def get_restaurant_info():
    restaurant_data = []
    address_data = []
    data = []
    load_wb = openpyxl.load_workbook('../DB_cell/restaurant_data.xlsx', data_only=True)
    load_ws = load_wb['restaurant']
    for row in load_ws.iter_rows(min_row=2):
        restaurant_data.append([row[1].value,row[3].value,row[0].value]) #식당명, 주소Id,

    load_wb2 = openpyxl.load_workbook('../DB_cell/address_data.xlsx', data_only=True)
    load_ws2 = load_wb2['address']
    for row in load_ws2.iter_rows(min_row=2):
        address_data.append([row[0].value,row[2].value]) #주소Id, district

    for r in restaurant_data:
        for a in address_data:
            if r[1] == a[0]:
                if a[1] == '관악구':
                    data.append([r[2],'봉천동 ' + r[0]])
                    break
                elif a[1] == '서대문구':
                    data.append([r[2],'창천동 ' + r[0]])
                    break
                elif a[1] == '동작구':
                    data.append([r[2],'상도동 ' + r[0]])
                    break
    #pprint(data)
    return data

def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]

def save_images(images, save_path):
    try:
        for index, image in enumerate(images[:10]):
            print(index)
            src = image.get_attribute('src')
            t = urlopen(src).read()
            file = open(os.path.join(save_path, str(index + 1) + ".jpg"),"wb")
            file.write(t)
            print("img save " + save_path + str(index + 1) + ".jpg")
    except Exception as e2:
        print('tt')

def create_folder_if_not_exists(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)

def crawl_restaurant_image(restaurants):
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome('./101/chromedriver',options=options)

    for r in restaurants:
        try:
            driver.get("https://naver.com")
            driver.implicitly_wait(time_to_wait=2)
            search_box = driver.find_element(by=By.ID, value = "query")
            search_box.send_keys(r[1])
            search_box.send_keys(Keys.ENTER)
            driver.implicitly_wait(time_to_wait=2)

            x=0

            try:
                driver.find_element(by=By.CLASS_NAME, value = "XNxh9").click()
                driver.switch_to.window(driver.window_handles[-1])
                driver.switch_to.frame("entryIframe")
                x+=1
            except Exception as e2:
                pass

            try:
                driver.find_element(by=By.CLASS_NAME, value = "_2opOK").click()
                driver.switch_to.window(driver.window_handles[-1])
                driver.switch_to.frame("entryIframe")
                x+=1
            except Exception as e2:
                pass

            try:
                img_type1 = driver.find_element(by='xpath',value = '//*[@id="app-root"]/div/div/div/div[7]/div[2]/div/div/div/div/div[1]/a')
                t = img_type1.text
                img_type1.click()

                save_path = "../restaurant_image/" + str(r[0]) + '_' + r[1] + "/" + t + "/"
                images = driver.find_elements(by=By.CLASS_NAME, value = "_img")
                create_folder_if_not_exists(save_path)
                save_images(images,save_path)
            except Exception as e2:
                pass

            try:
                img_type2 = driver.find_element(by='xpath',value = '//*[@id="app-root"]/div/div/div/div[7]/div[2]/div/div/div/div/div[2]/a')
                t = img_type2.text
                img_type2.click()

                save_path = "../restaurant_image/" + str(r[0]) + '_' + r[1] + "/" + t + "/"
                images = driver.find_elements(by=By.CLASS_NAME, value = "_img")
                create_folder_if_not_exists(save_path)
                save_images(images,save_path)
            except Exception as e2:
                pass


            try:
                img_type3 = driver.find_element(by='xpath',value = '//*[@id="app-root"]/div/div/div/div[7]/div[2]/div/div/div/div/div[3]/a')
                t = img_type3.text
                img_type3.click()

                save_path = "../restaurant_image/" + str(r[0]) + '_' + r[1] + "/" + t + "/"
                images = driver.find_elements(by=By.CLASS_NAME, value = "_img")
                create_folder_if_not_exists(save_path)
                save_images(images,save_path)
            except Exception as e2:
                pass

            try:
                img_type4 = driver.find_element(by='xpath',value = '//*[@id="app-root"]/div/div/div/div[7]/div[2]/div/div/div/div/div[4]/a')
                t = img_type4.text
                img_type4.click()

                save_path = "../restaurant_image/" + str(r[0]) + '_' + r[1] + "/" + t + "/"
                images = driver.find_elements(by=By.CLASS_NAME, value = "_img")
                create_folder_if_not_exists(save_path)
                save_images(images,save_path)
            except Exception as e2:
                pass

            try:
                img_type5 = driver.find_element(by='xpath',value = '//*[@id="app-root"]/div/div/div/div[7]/div[2]/div/div/div/div/div[5]/a')
                t = img_type5.text
                img_type5.click()

                save_path = "../restaurant_image/" + str(r[0]) + '_' + r[1] + "/" + t + "/"
                images = driver.find_elements(by=By.CLASS_NAME, value = "_img")
                create_folder_if_not_exists(save_path)
                save_images(images,save_path)
            except Exception as e2:
                pass
            if x==1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0]);

        except Exception as e1:
            #print(e1)
            print(r + " 정보 없음")
            driver.close()
            driver.switch_to.window(driver.window_handles[0]);
            pass

if __name__ == '__main__':
    restaurants = get_restaurant_info()
    restaurants_chunked = list_chunk(restaurants[2490:],7)
    start_time = time.time()
    f_list = ["서대문구레스토랑1.xlsx","서대문구레스토랑2.xlsx","서대문구레스토랑3.xlsx"]
    pool = multiprocessing.Pool(processes=3)
    pool.map(crawl_restaurant_image,restaurants_chunked)
    pool.close()
    pool.join()
    print(time.time()-start_time)
    #crawl_restaurant_image(restaurants)
